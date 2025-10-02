import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import time
import re
from urllib.parse import quote, urljoin
import json


class JazzShopParser:
    def __init__(self):
        self.base_url = "https://jazz-shop.ru"
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
            'Accept-Language': 'ru-RU,ru;q=0.9,en;q=0.8',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
        })

    def search_products(self, query):
        """Поиск товаров по запросу"""
        # Пробуем разные варианты поисковых URL
        search_urls = [
            f"{self.base_url}/search/?query={quote(query)}",
            f"{self.base_url}/search/?q={quote(query)}",
            f"{self.base_url}/search?query={quote(query)}",
            f"{self.base_url}/search?q={quote(query)}",
            f"{self.base_url}/catalog/search/?query={quote(query)}",
        ]

        products = []

        for search_url in search_urls:
            print(f"Пробуем URL: {search_url}")
            try:
                response = self.session.get(search_url, timeout=15)
                response.raise_for_status()

                # Пробуем разные кодировки
                for encoding in ['utf-8', 'windows-1251', 'cp1251']:
                    try:
                        response.encoding = encoding
                        soup = BeautifulSoup(response.text, 'html.parser')
                        break
                    except:
                        continue

                # Ищем товары разными способами
                products = self._find_products(soup)
                if products:
                    print(f"Найдено товаров: {len(products)}")
                    break
                else:
                    print("Товары не найдены, пробуем следующий URL...")

            except Exception as e:
                print(f"Ошибка при запросе {search_url}: {e}")
                continue

        # Если не нашли через поиск, пробуем парсить главную страницу или категории
        if not products:
            products = self._fallback_search(query)

        return products

    def _find_products(self, soup):
        """Поиск товаров на странице разными методами"""
        products = []

        # Метод 1: Ищем по классам
        class_selectors = [
            '.product',
            '.item',
            '.goods',
            '.catalog-item',
            '.product-item',
            '.item-product',
            '.card',
            '.product-card',
            '.shop-item'
        ]

        for selector in class_selectors:
            elements = soup.select(selector)
            if elements:
                print(f"Найдено элементов с селектором {selector}: {len(elements)}")
                for element in elements:
                    product = self._parse_product_element(element)
                    if product and product['Название'] != "Название не найдено":
                        products.append(product)
                if products:
                    return products

        # Метод 2: Ищем по структуре - элементы с ценами
        price_elements = soup.find_all(text=re.compile(r'руб|р\.|₽|цена', re.IGNORECASE))
        for price_element in price_elements:
            parent = price_element.parent
            product = self._parse_nearby_elements(parent)
            if product and product['Название'] != "Название не найдено":
                products.append(product)

        # Метод 3: Ищем все карточки товаров
        cards = soup.find_all(['div', 'article', 'li'], class_=True)
        for card in cards:
            card_classes = card.get('class', [])
            if any(cls for cls in card_classes if
                   any(word in cls.lower() for word in ['product', 'item', 'card', 'goods', 'shop'])):
                product = self._parse_product_element(card)
                if product and product['Название'] != "Название не найдено":
                    products.append(product)

        return products[:20]  # Ограничиваем количество

    def _parse_product_element(self, element):
        """Парсинг элемента товара"""
        try:
            product = {}

            # Название
            name = self._extract_name(element)
            product['Название'] = name if name else "Название не найдено"

            # Цена
            price = self._extract_price_from_element(element)
            product['Цена'] = price if price else "Цена не найдена"

            # Ссылка
            link = self._extract_link(element)
            product['Ссылка'] = link if link else ""

            # Бренд
            brand = self._extract_brand(element)
            product['Бренд'] = brand if brand else "Бренд не указан"

            # Наличие
            availability = self._extract_availability(element)
            product['Наличие'] = availability

            # Артикул
            article = self._extract_article(element)
            product['Артикул'] = article

            return product

        except Exception as e:
            print(f"Ошибка парсинга элемента: {e}")
            return None

    def _extract_name(self, element):
        """Извлечение названия товара"""
        name_selectors = [
            '.name',
            '.title',
            '.product-name',
            '.item-name',
            'h1', 'h2', 'h3', 'h4',
            'a[class*="name"]',
            'a[class*="title"]'
        ]

        for selector in name_selectors:
            name_elem = element.select_one(selector)
            if name_elem:
                text = name_elem.get_text(strip=True)
                if text and len(text) > 3:
                    return text

        # Ищем в ссылках
        links = element.find_all('a', href=True)
        for link in links:
            text = link.get_text(strip=True)
            if text and len(text) > 3 and not re.match(r'^(купить|в корзину|подробнее)$', text, re.IGNORECASE):
                return text

        return None

    def _extract_price_from_element(self, element):
        """Извлечение цены из элемента"""
        # Ищем цену в тексте элемента и его детей
        text = element.get_text()
        price_patterns = [
            r'(\d{1,3}(?:\s?\d{3})*(?:\,\d{2})?)\s*руб',
            r'(\d{1,3}(?:\s?\d{3})*(?:\,\d{2})?)\s*р\.',
            r'цена[:\s]*(\d{1,3}(?:\s?\d{3})*(?:\,\d{2})?)',
            r'₽\s*(\d{1,3}(?:\s?\d{3})*(?:\,\d{2})?)'
        ]

        for pattern in price_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                price_clean = matches[0].replace(' ', '').replace(',', '.')
                try:
                    return f"{float(price_clean):.2f}"
                except:
                    continue

        return None

    def _extract_link(self, element):
        """Извлечение ссылки на товар"""
        # Ищем ссылки в элементе
        links = element.find_all('a', href=True)
        for link in links:
            href = link.get('href')
            if href and not any(x in href.lower() for x in ['javascript:', '#', 'mailto:', 'tel:']):
                if href.startswith('/'):
                    return urljoin(self.base_url, href)
                elif href.startswith('http'):
                    return href
                else:
                    return urljoin(self.base_url, '/' + href.lstrip('/'))
        return None

    def _extract_brand(self, element):
        """Извлечение бренда"""
        brand_selectors = [
            '.brand',
            '.vendor',
            '.producer',
            '.manufacturer'
        ]

        for selector in brand_selectors:
            brand_elem = element.select_one(selector)
            if brand_elem:
                return brand_elem.get_text(strip=True)

        return None

    def _extract_availability(self, element):
        """Извлечение информации о наличии"""
        text = element.get_text().lower()
        if any(word in text for word in ['в наличии', 'есть', 'доступен', 'available', 'купить']):
            return "В наличии"
        elif any(word in text for word in ['нет в наличии', 'распродан', 'ожидается', 'под заказ']):
            return "Нет в наличии"
        else:
            return "Неизвестно"

    def _extract_article(self, element):
        """Извлечение артикула"""
        text = element.get_text()
        patterns = [
            r'арт[.\s]*[:#]?\s*([a-zA-Z0-9-]+)',
            r'art[.\s]*[:#]?\s*([a-zA-Z0-9-]+)',
            r'код[.\s]*[:#]?\s*([a-zA-Z0-9-]+)',
            r'code[.\s]*[:#]?\s*([a-zA-Z0-9-]+)'
        ]

        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return match.group(1)

        return "Артикул не найден"

    def _parse_nearby_elements(self, element):
        """Парсинг соседних элементов вокруг цены"""
        # Ищем название в родительских элементах
        for i in range(3):
            parent = element
            for j in range(i):
                parent = getattr(parent, 'parent', None)
                if not parent:
                    break

            if parent:
                return self._parse_product_element(parent)

        return None

    def _fallback_search(self, query):
        """Резервный метод поиска"""
        print("Используем резервный метод поиска...")

        # Пробуем получить главную страницу и найти категории
        try:
            response = self.session.get(self.base_url, timeout=15)
            soup = BeautifulSoup(response.text, 'html.parser')

            # Ищем ссылки, содержащие запрос
            query_words = query.lower().split()
            links = soup.find_all('a', href=True, string=True)

            products = []
            for link in links:
                link_text = link.get_text().lower()
                if any(word in link_text for word in query_words):
                    product = {
                        'Название': link.get_text(strip=True),
                        'Цена': "Неизвестно",
                        'Ссылка': urljoin(self.base_url, link['href']),
                        'Бренд': "Не указан",
                        'Наличие': "Неизвестно",
                        'Артикул': "Не найден"
                    }
                    products.append(product)

            return products[:10]

        except Exception as e:
            print(f"Ошибка в резервном поиске: {e}")
            return []

    def get_detailed_info(self, product_url):
        """Получение детальной информации о товаре"""
        if not product_url:
            return {}

        try:
            response = self.session.get(product_url, timeout=15)
            soup = BeautifulSoup(response.text, 'html.parser')

            detailed_info = {}

            # Описание
            description = self._extract_description(soup)
            detailed_info['Описание'] = description

            # Характеристики
            characteristics = self._extract_characteristics(soup)
            detailed_info['Характеристики'] = json.dumps(characteristics, ensure_ascii=False)

            return detailed_info

        except Exception as e:
            print(f"Ошибка получения детальной информации: {e}")
            return {
                'Описание': 'Не удалось получить описание',
                'Характеристики': '{}'
            }

    def _extract_description(self, soup):
        """Извлечение описания товара"""
        desc_selectors = [
            '.description',
            '.product-description',
            '.item-description',
            '[class*="desc"]',
            '#description'
        ]

        for selector in desc_selectors:
            elem = soup.select_one(selector)
            if elem:
                text = elem.get_text(strip=True)
                if text and len(text) > 10:
                    return text[:500] + "..." if len(text) > 500 else text

        return "Описание не найдено"

    def _extract_characteristics(self, soup):
        """Извлечение характеристик товара"""
        specs = {}

        # Ищем таблицы характеристик
        tables = soup.find_all('table')
        for table in tables:
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all(['td', 'th'])
                if len(cells) >= 2:
                    key = cells[0].get_text(strip=True)
                    value = cells[1].get_text(strip=True)
                    if key and value:
                        specs[key] = value

        # Ищем списки характеристик
        lists = soup.find_all(['ul', 'ol'])
        for list_elem in lists:
            items = list_elem.find_all('li')
            for item in items:
                text = item.get_text(strip=True)
                if ':' in text:
                    key, value = text.split(':', 1)
                    specs[key.strip()] = value.strip()

        return specs

    def save_to_excel(self, products, filename='jazz_shop_products.xlsx'):
        """Сохранение результатов в Excel"""
        if not products:
            print("Нет данных для сохранения")
            return False

        try:
            wb = Workbook()
            ws_products = wb.active
            ws_products.title = "Товары"

            # Заголовки
            headers = ['Название', 'Бренд', 'Цена', 'Наличие', 'Артикул', 'Ссылка', 'Описание', 'Характеристики']
            for col, header in enumerate(headers, 1):
                cell = ws_products.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            # Данные
            for row, product in enumerate(products, 2):
                ws_products.cell(row=row, column=1, value=product.get('Название', ''))
                ws_products.cell(row=row, column=2, value=product.get('Бренд', ''))
                ws_products.cell(row=row, column=3, value=product.get('Цена', ''))
                ws_products.cell(row=row, column=4, value=product.get('Наличие', ''))
                ws_products.cell(row=row, column=5, value=product.get('Артикул', ''))
                ws_products.cell(row=row, column=6, value=product.get('Ссылка', ''))
                ws_products.cell(row=row, column=7, value=product.get('Описание', ''))
                ws_products.cell(row=row, column=8, value=product.get('Характеристики', ''))

            # Авто-ширина колонок
            for column in ws_products.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_products.column_dimensions[column_letter].width = adjusted_width

            # Сохраняем файл
            wb.save(filename)
            print(f"Файл сохранен: {filename}")
            return True

        except Exception as e:
            print(f"Ошибка сохранения в Excel: {e}")
            return False